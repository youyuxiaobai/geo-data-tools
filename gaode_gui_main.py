#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
高德地图 POI 爬虫 - 桌面版
Tkinter GUI + PostgreSQL 存储，支持 PyInstaller 打包为单文件 exe
"""

import json
import os
import sys
import time
import queue
import threading
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime

import requests
import psycopg2
import shapefile as shp
from shapely.geometry import Point, Polygon, MultiPolygon

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ============================================================
# 路径工具
# ============================================================

def _get_app_dir():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

def _get_data_dir():
    if getattr(sys, 'frozen', False):
        return os.path.join(sys._MEIPASS, "Gaode_poi")
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "Gaode_poi")

def _resolve_data_path(filename):
    candidates = [
        os.path.join(_get_app_dir(), filename),
        os.path.join(_get_data_dir(), filename),
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "Gaode_poi", filename),
    ]
    for p in candidates:
        if os.path.exists(p):
            return p
    return candidates[-1]

# ============================================================
# 常量
# ============================================================

FIELDS = {
    "编号":     "id",
    "行业类型": "biz_type",
    "名称":     "name",
    "类别":     "type",
    "地址":     "address",
    "电话":     "tel",
    "经纬度":   "location",
    "省编码":   "pcode",
    "省份":     "pname",
    "城市编码": "citycode",
    "城市":     "cityname",
    "区县编码": "adcode",
    "区县":     "adname",
    "商圈":     "business_area",
    "抓取批次": "batch_id",
}

TABLE_COLUMNS_DEF = (
    "编号 VARCHAR(64) PRIMARY KEY, "
    "行业类型 TEXT, 名称 TEXT, 类别 TEXT, 地址 TEXT, "
    "电话 TEXT, 经纬度 TEXT, 省编码 TEXT, 省份 TEXT, "
    "城市编码 TEXT, 城市 TEXT, 区县编码 TEXT, 区县 TEXT, 商圈 TEXT, "
    "抓取批次 TEXT"
)

DEFAULT_CONFIG = {
    "api_key": "",
    "poi_types": "061205",
    "page_size": 20,
    "db_host": "localhost",
    "db_port": 5432,
    "db_user": "postgres",
    "db_password": "",
    "db_name": "pgadmind",
    "table_name": "gaode_poi",
}

# ============================================================
# 数据加载
# ============================================================

def _safe_read_json(path):
    """读取 JSON 文件，自动尝试 UTF-8 / UTF-8-SIG / GBK 编码"""
    for enc in ("utf-8-sig", "utf-8", "gbk", "gb2312"):
        try:
            with open(path, encoding=enc) as f:
                return json.load(f)
        except (UnicodeDecodeError, json.JSONDecodeError):
            continue
    # 最终尝试原始字节
    with open(path, "rb") as f:
        return json.loads(f.read())


def load_cities():
    path = _resolve_data_path("city.json")
    return _safe_read_json(path)

def build_poi_tree():
    """解析 amap_poicode.xlsx，构建 大类→中类→小类 三级树"""
    if not HAS_OPENPYXL:
        return {}
    path = _resolve_data_path("amap_poicode.xlsx")
    if not os.path.exists(path):
        return {}
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_names = wb.sheetnames
    data_sheet = sheet_names[-1] if len(sheet_names) >= 3 else sheet_names[0]
    ws = wb[data_sheet]

    tree = {}  # {big_code: {"name": str, "children": {mid_code: {...}}}}
    # 一级缓存：大类名称
    big_names = {}
    # 二级缓存：中类名称
    mid_names = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        code = str(row[1]).strip() if row[1] else ""
        big_name = str(row[2]).strip() if row[2] else ""
        mid_name = str(row[3]).strip() if row[3] else ""
        sub_name = str(row[4]).strip() if row[4] else ""

        if not code or not code.isdigit():
            continue

        # 大类 (XX0000)
        if code.endswith("0000"):
            big_names[code] = big_name
            if code not in tree:
                tree[code] = {"name": big_name, "children": {}}
            continue

        # 中类 (XXYY00)
        if code.endswith("00"):
            big_code = code[:2] + "0000"
            mid_names[code] = mid_name
            if big_code not in tree:
                tree[big_code] = {"name": big_name, "children": {}}
            if code not in tree[big_code]["children"]:
                tree[big_code]["children"][code] = {"name": mid_name, "children": {}}
            continue

        # 小类 (其余)
        big_code = code[:2] + "0000"
        mid_code = code[:4] + "00"
        if big_code not in tree:
            tree[big_code] = {"name": big_name, "children": {}}
        if mid_code not in tree[big_code]["children"]:
            tree[big_code]["children"][mid_code] = {
                "name": mid_names.get(mid_code, mid_name), "children": {}
            }
        tree[big_code]["children"][mid_code]["children"][code] = {
            "name": sub_name, "children": {}
        }

    wb.close()
    return tree

def build_city_tree(cities):
    """从扁平城市列表构建 省→市→区县 三级树"""
    tree = {}

    # 省级 (adcode 以 0000 结尾)
    for c in cities:
        adcode = c["adcode"]
        if adcode.endswith("0000"):
            key = adcode[:2]
            tree[key] = {"name": c["city"], "adcode": adcode, "children": {}}

    # 市级 (adcode 以 00 结尾但不是 0000)
    for c in cities:
        adcode = c["adcode"]
        if adcode.endswith("00") and not adcode.endswith("0000"):
            prov_key = adcode[:2]
            if prov_key in tree:
                tree[prov_key]["children"][adcode] = {
                    "name": c["city"], "adcode": adcode, "children": {}
                }

    # 区县级 (其余)
    for c in cities:
        adcode = c["adcode"]
        if adcode.endswith("00"):
            continue
        prov_key = adcode[:2]
        city_key = adcode[:4] + "00"
        if prov_key not in tree:
            continue
        if city_key in tree[prov_key]["children"]:
            tree[prov_key]["children"][city_key]["children"][adcode] = {
                "name": c["city"], "adcode": adcode, "children": {}
            }
        else:
            # 区县直属省（直辖市）
            tree[prov_key]["children"][adcode] = {
                "name": c["city"], "adcode": adcode, "children": {}
            }

    return tree

# ============================================================
# ConfigManager
# ============================================================

class ConfigManager:
    def __init__(self):
        self.config_path = os.path.join(_get_app_dir(), "config.json")
        self.config = dict(DEFAULT_CONFIG)
        self.load()

    def load(self):
        try:
            with open(self.config_path, encoding="utf-8") as f:
                loaded = json.load(f)
                self.config.update(loaded)
        except (FileNotFoundError, json.JSONDecodeError, UnicodeDecodeError):
            # 配置文件损坏或编码错误，删除重建
            try:
                os.remove(self.config_path)
            except OSError:
                pass

    def save(self):
        with open(self.config_path, "w", encoding="utf-8") as f:
            json.dump(self.config, f, ensure_ascii=False, indent=2)

    def get(self, key, default=None):
        return self.config.get(key, default)

# ============================================================
# CrawlerEngine (后台线程)
# ============================================================

class CrawlerEngine(threading.Thread):
    def __init__(self, config, cities, log_queue, stop_event, batch_id):
        super().__init__(daemon=True)
        self.config = config
        self.cities = cities
        self.log_queue = log_queue
        self.stop_event = stop_event
        self.batch_id = batch_id

    def _log(self, msg):
        self.log_queue.put(("log", msg))

    def run(self):
        try:
            self._do_run()
        except Exception as e:
            self._log(f"严重错误: {e}")

    def _do_run(self):
        cfg = self.config
        self._log("正在连接数据库...")
        try:
            conn = psycopg2.connect(
                host=cfg["db_host"],
                port=cfg["db_port"],
                user=cfg["db_user"],
                password=cfg["db_password"],
                dbname=cfg["db_name"],
                client_encoding="UTF8",
            )
        except Exception as e:
            self._log(f"数据库连接失败: {e}")
            self.log_queue.put(("error", str(e)))
            return

        try:
            self._ensure_table(conn)
            columns = list(FIELDS.keys())
            col_list = ", ".join(f'"{c}"' for c in columns)
            placeholders = ", ".join(f"%({c})s" for c in columns)
            tbl = cfg["table_name"]

            insert_sql = f'INSERT INTO "{tbl}" ({col_list}) VALUES ({placeholders})'
            check_sql = f'SELECT 1 FROM "{tbl}" WHERE "编号" = %(编号)s'

            api_key = cfg["api_key"]
            poi_types = cfg["poi_types"]
            page_size = int(cfg.get("page_size", 20))
            total_cities = len(self.cities)
            overall_inserted = 0

            for city_idx, city in enumerate(self.cities):
                if self.stop_event.is_set():
                    self._log("用户停止爬取")
                    break

                adcode = city["adcode"]
                city_name = city["city"]
                self._log(f"[{city_idx + 1}/{total_cities}] 开始: {city_name} ({adcode})")
                self.log_queue.put(("progress_city", city_idx, total_cities, city_name))

                url = (
                    "http://restapi.amap.com/v3/place/text"
                    f"?key={api_key}"
                    f"&types={poi_types}"
                    f"&city={adcode}"
                    "&citylimit=true"
                    "&children=1"
                    f"&offset={page_size}"
                    "&page=page_index"
                    "&extensions=all"
                )

                time.sleep(0.5)
                try:
                    resp = requests.get(url.replace("page_index", "1"), timeout=30)
                    data = json.loads(resp.content)
                except Exception as e:
                    self._log(f"  请求失败: {e}")
                    continue

                total_count = int(data.get("count", 0))
                if total_count == 0:
                    self._log("  无数据，跳过")
                    continue

                total_pages = (total_count + page_size - 1) // page_size
                self._log(f"  共 {total_count} 条, {total_pages} 页")

                city_inserted = 0
                for page in range(1, total_pages + 1):
                    if self.stop_event.is_set():
                        break

                    if page > 1:
                        time.sleep(0.5)

                    try:
                        resp = requests.get(url.replace("page_index", str(page)), timeout=30)
                        poi_json = json.loads(resp.content)
                    except Exception as e:
                        self._log(f"  第{page}页请求失败: {e}")
                        continue

                    pois = poi_json.get("pois")
                    if not pois:
                        continue

                    page_inserted = 0
                    with conn.cursor() as cur:
                        for poi in pois:
                            row = {cn: poi.get(api_field) for cn, api_field in FIELDS.items()}
                            row["抓取批次"] = self.batch_id
                            try:
                                cur.execute(check_sql, {"编号": row["编号"]})
                                if cur.fetchone() is None:
                                    cur.execute(insert_sql, row)
                                    page_inserted += 1
                            except Exception as e:
                                self._log(f"  写入失败: {e}")

                    conn.commit()
                    city_inserted += page_inserted
                    self._log(f"  第{page}/{total_pages}页: 写入 {page_inserted} 条")
                    self.log_queue.put(("progress_page", page, total_pages))

                overall_inserted += city_inserted
                self._log(f"  {city_name} 完成: 共写入 {city_inserted} 条")

            self._log(f"全部完成！共写入 {overall_inserted} 条数据")
            self.log_queue.put(("done", overall_inserted))

        finally:
            conn.close()
            self._log("数据库连接已关闭")

    def _ensure_table(self, conn):
        tbl = self.config["table_name"]
        sql = f'CREATE TABLE IF NOT EXISTS "{tbl}" ({TABLE_COLUMNS_DEF})'
        with conn.cursor() as cur:
            cur.execute(sql)
        conn.commit()
        self._log(f"数据表 \"{tbl}\" 已就绪")

# ============================================================
# POI 类型树形选择对话框
# ============================================================

class PoiTypeDialog(tk.Toplevel):
    def __init__(self, parent, poi_tree, current_code):
        super().__init__(parent)
        self.title("选择 POI 类别 — 双击选中")
        self.geometry("650x550")
        self.transient(parent)
        self.grab_set()
        self.result = None
        self.poi_tree = poi_tree

        # 搜索框
        search_frame = ttk.Frame(self, padding="5")
        search_frame.pack(fill=tk.X)
        ttk.Label(search_frame, text="搜索:").pack(side=tk.LEFT)
        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", lambda *a: self._apply_filter())
        search_entry = ttk.Entry(search_frame, textvariable=self.search_var, width=30)
        search_entry.pack(side=tk.LEFT, padx=(5, 0))
        ttk.Label(search_frame, text="（输入大类/中类/小类名称或编码）",
                  foreground="gray").pack(side=tk.LEFT, padx=(5, 0))

        # 树形控件
        tree_frame = ttk.Frame(self, padding="5")
        tree_frame.pack(fill=tk.BOTH, expand=True)

        self.tree = ttk.Treeview(tree_frame, columns=("code",), show="tree headings",
                                 selectmode="browse")
        self.tree.heading("#0", text="POI 分类")
        self.tree.heading("code", text="编码")
        self.tree.column("#0", width=420)
        self.tree.column("code", width=100, anchor=tk.CENTER)

        tree_scroll = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=tree_scroll.set)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        tree_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        # 存储每个 item 的代码
        self._item_codes = {}

        # 填充
        self._populate(poi_tree)

        # 展开第一层
        for child in self.tree.get_children():
            self.tree.item(child, open=True)

        # 按钮
        btn_frame = ttk.Frame(self, padding="5")
        btn_frame.pack(fill=tk.X)
        ttk.Label(btn_frame, text="提示：双击任意节点选择其编码，API 支持任意层级编码",
                  foreground="gray").pack(side=tk.LEFT)
        ttk.Button(btn_frame, text="确定", command=self._on_ok).pack(side=tk.RIGHT, padx=(5, 0))
        ttk.Button(btn_frame, text="取消", command=self.destroy).pack(side=tk.RIGHT)

        self.tree.bind("<Double-1>", lambda e: self._on_ok())

        # 若当前编码匹配，尝试展开定位
        if current_code:
            self._locate_code(current_code)

        search_entry.focus()
        self.wait_window()

    def _populate(self, tree, parent="", filter_keyword=""):
        self.tree.delete(*self.tree.get_children())
        self._item_codes.clear()
        self._insert_nodes(tree, parent, filter_keyword)

    def _insert_nodes(self, nodes, parent, keyword):
        """递归插入节点。keyword 为空则全部插入，否则只插入匹配的子树"""
        for code in sorted(nodes.keys()):
            node = nodes[code]
            name = node["name"]
            children = node.get("children", {})

            # 检查本节点或子节点是否匹配
            match_self = (not keyword or
                          keyword in code.lower() or
                          keyword in name.lower())
            match_children = self._any_child_match(children, keyword) if keyword else True

            if not match_self and not match_children:
                continue

            display = name
            item_id = self.tree.insert(parent, tk.END, text=display, values=(code,),
                                       open=(keyword != "" and match_children))
            self._item_codes[item_id] = code

            if children:
                self._insert_nodes(children, item_id, keyword)

    def _any_child_match(self, nodes, keyword):
        if not keyword:
            return True
        for code, node in nodes.items():
            if (keyword in code.lower() or keyword in node["name"].lower() or
                    self._any_child_match(node.get("children", {}), keyword)):
                return True
        return False

    def _apply_filter(self):
        keyword = self.search_var.get().strip().lower()
        self._populate(self.poi_tree, filter_keyword=keyword)
        # 展开匹配结果
        if keyword:
            self._expand_all()

    def _expand_all(self, parent=""):
        for child in self.tree.get_children(parent):
            self.tree.item(child, open=True)
            self._expand_all(child)

    def _locate_code(self, target):
        """展开到包含目标编码的节点"""
        for item_id, code in self._item_codes.items():
            if code == target:
                self.tree.see(item_id)
                self.tree.selection_set(item_id)
                # 展开父节点路径
                parent = self.tree.parent(item_id)
                while parent:
                    self.tree.item(parent, open=True)
                    parent = self.tree.parent(parent)
                return

    def _on_ok(self):
        sel = self.tree.selection()
        if sel:
            self.result = self._item_codes.get(sel[0], "")
            self.destroy()

# ============================================================
# 主 GUI 类
# ============================================================

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("高德地图 POI 爬虫工具")
        self.geometry("1050x750")
        self.minsize(900, 650)

        self.config_mgr = ConfigManager()
        self.cities = []
        self.city_tree_data = {}
        self.poi_tree = {}
        self.stop_event = threading.Event()
        self.crawler = None
        self.log_queue = queue.Queue()
        self._selected_adcodes = set()
        self._all_adcodes = []
        self.current_batch_id = None

        # 加载数据
        try:
            self.cities = load_cities()
            self._all_adcodes = [c["adcode"] for c in self.cities]
            saved = self.config_mgr.get("selected_cities", None)
            if saved is None:
                self._selected_adcodes = set(self._all_adcodes)
            else:
                self._selected_adcodes = set(saved)
            self.city_tree_data = build_city_tree(self.cities)
        except Exception as e:
            messagebox.showwarning("警告", f"加载城市列表失败:\n{e}")

        try:
            self.poi_tree = build_poi_tree()
        except Exception as e:
            messagebox.showwarning("警告", f"加载 POI 分类表失败:\n{e}\n\n"
                                   "POI 浏览功能不可用，可手动输入编码。")

        self._apply_styles()
        self._build_ui()
        self._load_config_to_ui()
        self._poll_log()
        self.protocol("WM_DELETE_WINDOW", self._on_close)

    def _apply_styles(self):
        style = ttk.Style()

        BG = "#F8F6FF"
        SURFACE = "#FFFFFF"
        TEXT = "#1E1B4B"
        TEXT_SEC = "#6D63B8"
        ACCENT = "#7C3AED"

        self.configure(bg="#0A0A1A")

        style.configure(".", font=("Microsoft YaHei UI", 9))
        style.configure("TFrame", background=BG)
        style.configure("TLabelframe", background=SURFACE)
        style.configure("TLabelframe.Label", background=SURFACE, foreground=TEXT,
                         font=("Microsoft YaHei UI", 10, "bold"))
        style.configure("TNotebook", background=BG, borderwidth=0)
        style.configure("TNotebook.Tab", padding=(18, 8), font=("Microsoft YaHei UI", 10),
                         background="#F1EEFC", foreground=TEXT_SEC)
        style.map("TNotebook.Tab",
                  background=[("selected", SURFACE)],
                  foreground=[("selected", ACCENT)])
        style.configure("TLabel", background=BG, foreground=TEXT)
        style.configure("TEntry", padding=6, fieldbackground="#FAFAFE")
        style.configure("TButton", padding=(14, 6), font=("Microsoft YaHei UI", 9))
        style.configure("TSpinbox", padding=4)
        style.configure("Treeview", rowheight=28, background=SURFACE,
                         fieldbackground=SURFACE)
        style.configure("Treeview.Heading", background="#E8E4F8", foreground=TEXT,
                         font=("Microsoft YaHei UI", 9, "bold"), padding=(8, 5))
        style.map("Treeview",
                  background=[("selected", "#7C3AED")],
                  foreground=[("selected", "white")])
        style.configure("TProgressbar", thickness=14, troughcolor="#E8E4F8",
                         background="#7C3AED")
        self.option_add("*Text.font", ("Consolas", 9))
        self.option_add("*Text.background", "#FFFFFF")
        self.option_add("*Text.foreground", "#1E1B4B")
        self.option_add("*Text.selectBackground", "#E8E4F8")
        self.option_add("*Text.background", "#1E293B")
        self.option_add("*Text.foreground", "#E2E8F0")

    # ---------- UI 构建 ----------

    def _build_ui(self):
        # 顶部炫彩标题栏
        header = tk.Frame(self, bg="#0A0A1A", height=48)
        header.pack(fill=tk.X)
        header.pack_propagate(False)

        # 左侧标题
        title_frame = tk.Frame(header, bg="#0A0A1A")
        title_frame.pack(side=tk.LEFT, padx=15, pady=6)
        tk.Label(title_frame, text="✦", fg="#C084FC", bg="#0A0A1A",
                 font=("Microsoft YaHei UI", 14)).pack(side=tk.LEFT)
        tk.Label(title_frame, text=" 高德地图 POI 爬虫工具 ",
                 bg="#0A0A1A", fg="#E9D5FF",
                 font=("Microsoft YaHei UI", 13, "bold")).pack(side=tk.LEFT)
        tk.Label(title_frame, text="✦", fg="#C084FC", bg="#0A0A1A",
                 font=("Microsoft YaHei UI", 14)).pack(side=tk.LEFT)

        # 右侧版本 + 开发者
        info_frame = tk.Frame(header, bg="#0A0A1A")
        info_frame.pack(side=tk.RIGHT, padx=15, pady=8)
        tk.Label(info_frame, text="v2.5 Pro", fg="#A78BFA", bg="#0A0A1A",
                 font=("Consolas", 9, "bold")).pack(side=tk.RIGHT, padx=(15, 0))
        tk.Label(info_frame, text="白栋博  bdb15896810691", fg="#6C3CE0",
                 bg="#0A0A1A", font=("Microsoft YaHei UI", 9)).pack(side=tk.RIGHT)

        # 主容器 — 用 PanedWindow 确保缩放时各部分都可见
        main_frame = ttk.Frame(self, padding="8")
        main_frame.pack(fill=tk.BOTH, expand=True)

        pane = ttk.PanedWindow(main_frame, orient=tk.VERTICAL)
        pane.pack(fill=tk.BOTH, expand=True)

        # 上部分：标签页
        top_section = ttk.Frame(pane, height=360)
        self.notebook = ttk.Notebook(top_section)
        self.notebook.pack(fill=tk.BOTH, expand=True)

        self._build_api_tab()
        self._build_db_tab()
        self._build_city_tab()
        self._build_district_tab()

        # 开发者联系卡片
        contact_bar = tk.Frame(top_section, bg="#1A1040", height=32)
        contact_bar.pack(fill=tk.X, pady=(4, 0))
        contact_bar.pack_propagate(False)
        tk.Label(contact_bar, text="✦  如有问题请联系开发者：",
                 bg="#1A1040", fg="#C084FC",
                 font=("Microsoft YaHei UI", 9, "bold")).pack(side=tk.LEFT, padx=(15, 0), pady=5)
        tk.Label(contact_bar, text="白栋博",
                 bg="#1A1040", fg="#E9D5FF",
                 font=("Microsoft YaHei UI", 10, "bold")).pack(side=tk.LEFT, pady=5)
        tk.Label(contact_bar, text="  |  微信：bdb15896810691",
                 bg="#1A1040", fg="#A78BFA",
                 font=("Microsoft YaHei UI", 10)).pack(side=tk.LEFT, pady=5)
        tk.Label(contact_bar, text="  ✦",
                 bg="#1A1040", fg="#C084FC",
                 font=("Microsoft YaHei UI", 9)).pack(side=tk.RIGHT, padx=(0, 15), pady=5)

        pane.add(top_section, weight=5)

        # 下部分：控件 + 日志
        bottom_section = ttk.Frame(pane, height=280)

        # 装饰线
        sep = tk.Frame(bottom_section, bg="#7C3AED", height=2)
        sep.pack(fill=tk.X, pady=(0, 4))

        # 底部控制区
        ctrl = ttk.Frame(bottom_section)
        ctrl.pack(fill=tk.X, pady=(0, 2))

        # 进度 + 状态
        pframe = tk.Frame(ctrl, bg="#F8F6FF")
        pframe.pack(fill=tk.X, pady=(0, 2))
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(pframe, variable=self.progress_var, maximum=100)
        self.progress_bar.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 8))
        self.status_label = tk.Label(pframe, text="✦ 就绪，等待开始 ✦",
                                     font=("Microsoft YaHei UI", 9, "bold"),
                                     fg="#7C3AED", bg="#F8F6FF", anchor=tk.E)
        self.status_label.pack(side=tk.RIGHT)

        self.sub_progress_var = tk.DoubleVar()
        self.sub_progress_bar = ttk.Progressbar(ctrl, variable=self.sub_progress_var, maximum=100)
        self.sub_progress_bar.pack(fill=tk.X, pady=(0, 3))

        # 按钮
        btn_bar = tk.Frame(ctrl, bg="#F8F6FF")
        btn_bar.pack(fill=tk.X, pady=(0, 2))

        self.btn_start = tk.Button(btn_bar, text="▶  开 始 爬 取", command=self._on_start,
                                   bg="#7C3AED", fg="white", font=("Microsoft YaHei UI", 11, "bold"),
                                   padx=22, pady=5, borderwidth=0, cursor="hand2",
                                   activebackground="#6D28D9", activeforeground="white")
        self.btn_start.pack(side=tk.LEFT, padx=(0, 6))

        self.btn_stop = tk.Button(btn_bar, text="■ 停 止", command=self._on_stop,
                                  state=tk.DISABLED, bg="#DC2626", fg="white",
                                  font=("Microsoft YaHei UI", 11, "bold"),
                                  padx=18, pady=5, borderwidth=0, cursor="hand2",
                                  activebackground="#B91C1C", activeforeground="white",
                                  disabledforeground="#FCA5A5")
        self.btn_stop.pack(side=tk.LEFT, padx=(0, 15))

        ttk.Button(btn_bar, text="💾 保存配置", command=self._on_save_config).pack(
            side=tk.LEFT, padx=(0, 4))
        ttk.Button(btn_bar, text="🔗 测试连接", command=self._on_test_db).pack(
            side=tk.LEFT, padx=(0, 4))
        ttk.Button(btn_bar, text="📋 创建数据表", command=self._on_create_table).pack(
            side=tk.LEFT, padx=(0, 4))
        ttk.Button(btn_bar, text="📦 导出 Shapefile", command=self._on_export_shp).pack(
            side=tk.LEFT)

        # 日志 — 白色风格
        log_frame = tk.Frame(bottom_section, bg="#FFFFFF", highlightbackground="#E8E4F8",
                              highlightthickness=1)
        log_frame.pack(fill=tk.BOTH, expand=True)

        self.log_text = tk.Text(log_frame, wrap=tk.WORD, state=tk.DISABLED,
                                font=("Consolas", 9), bg="#FFFFFF", fg="#1E1B4B",
                                insertbackground="#7C3AED", selectbackground="#E8E4F8",
                                borderwidth=0, padx=8, pady=4)
        log_scroll = ttk.Scrollbar(log_frame, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=log_scroll.set)
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        log_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        pane.add(bottom_section, weight=3)

    def _build_api_tab(self):
        tab = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(tab, text="API 配置")

        row = 0
        ttk.Label(tab, text="高德 API Key:").grid(row=row, column=0, sticky=tk.W, pady=6)
        self.entry_api_key = ttk.Entry(tab, width=48)
        self.entry_api_key.grid(row=row, column=1, sticky=tk.EW, pady=6, padx=(10, 0))
        row += 1

        ttk.Label(tab, text="POI 类别编码:").grid(row=row, column=0, sticky=tk.W, pady=6)
        type_frame = ttk.Frame(tab)
        type_frame.grid(row=row, column=1, sticky=tk.EW, pady=6, padx=(10, 0))
        self.entry_poi_types = ttk.Entry(type_frame, width=14)
        self.entry_poi_types.pack(side=tk.LEFT, fill=tk.X, expand=True)
        ttk.Button(type_frame, text="浏览...", command=self._on_browse_poi).pack(
            side=tk.LEFT, padx=(5, 0))
        row += 1

        ttk.Label(tab, text="每页条数 (1-25):").grid(row=row, column=0, sticky=tk.W, pady=6)
        self.spin_page_size = ttk.Spinbox(tab, from_=1, to=25, width=8)
        self.spin_page_size.grid(row=row, column=1, sticky=tk.W, pady=6, padx=(10, 0))
        row += 1

        ttk.Label(tab, text='提示：点击「浏览」可从分类树中选择，支持按名称搜索（如"景区""商业"）',
                  foreground="gray").grid(row=row, column=0, columnspan=2, sticky=tk.W, pady=(10, 0))

        tab.columnconfigure(1, weight=1)

    def _build_db_tab(self):
        tab = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(tab, text="数据库配置")

        row = 0
        ttk.Label(tab, text="主机:").grid(row=row, column=0, sticky=tk.W, pady=5)
        self.entry_db_host = ttk.Entry(tab, width=22)
        self.entry_db_host.grid(row=row, column=1, sticky=tk.W, pady=5, padx=(10, 0))
        ttk.Label(tab, text="端口:").grid(row=row, column=2, sticky=tk.W, pady=5, padx=(20, 0))
        self.entry_db_port = ttk.Entry(tab, width=8)
        self.entry_db_port.grid(row=row, column=3, sticky=tk.W, pady=5, padx=(10, 0))
        row += 1

        ttk.Label(tab, text="用户名:").grid(row=row, column=0, sticky=tk.W, pady=5)
        self.entry_db_user = ttk.Entry(tab, width=22)
        self.entry_db_user.grid(row=row, column=1, sticky=tk.W, pady=5, padx=(10, 0))
        ttk.Label(tab, text="密码:").grid(row=row, column=2, sticky=tk.W, pady=5, padx=(20, 0))
        self.entry_db_password = ttk.Entry(tab, width=22, show="*")
        self.entry_db_password.grid(row=row, column=3, sticky=tk.W, pady=5, padx=(10, 0))
        row += 1

        ttk.Label(tab, text="数据库名:").grid(row=row, column=0, sticky=tk.W, pady=5)
        self.entry_db_name = ttk.Entry(tab, width=22)
        self.entry_db_name.grid(row=row, column=1, sticky=tk.W, pady=5, padx=(10, 0))
        ttk.Label(tab, text="表名:").grid(row=row, column=2, sticky=tk.W, pady=5, padx=(20, 0))
        self.entry_table_name = ttk.Entry(tab, width=22)
        self.entry_table_name.grid(row=row, column=3, sticky=tk.W, pady=5, padx=(10, 0))

    # ---------- 城市选择（树形 + 级联复选框） ----------

    def _build_city_tab(self):
        tab = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(tab, text="城市选择")

        # 顶部工具栏
        top = ttk.Frame(tab)
        top.pack(fill=tk.X, pady=(0, 5))

        ttk.Label(top, text="筛选:").pack(side=tk.LEFT)
        self.city_filter_var = tk.StringVar()
        self.city_filter_var.trace_add("write", lambda *a: self._apply_city_filter())
        ttk.Entry(top, textvariable=self.city_filter_var, width=18).pack(
            side=tk.LEFT, padx=(5, 15))

        ttk.Button(top, text="全选", command=self._city_select_all).pack(
            side=tk.LEFT, padx=(0, 3))
        ttk.Button(top, text="全不选", command=self._city_select_none).pack(
            side=tk.LEFT, padx=(0, 3))
        ttk.Button(top, text="反选", command=self._city_invert).pack(
            side=tk.LEFT, padx=(0, 3))
        ttk.Button(top, text="展开全部", command=self._city_expand_all).pack(
            side=tk.LEFT, padx=(0, 3))
        ttk.Button(top, text="折叠全部", command=self._city_collapse_all).pack(
            side=tk.LEFT)

        self.city_count_label = ttk.Label(top, text="")
        self.city_count_label.pack(side=tk.RIGHT)

        # 树形控件
        tree_frame = ttk.Frame(tab)
        tree_frame.pack(fill=tk.BOTH, expand=True)

        self.city_tree = ttk.Treeview(tree_frame, columns=("adcode",), show="tree headings",
                                      selectmode="extended")
        self.city_tree.heading("#0", text="行政区划")
        self.city_tree.heading("adcode", text="代码")
        self.city_tree.column("#0", width=320)
        self.city_tree.column("adcode", width=100, anchor=tk.CENTER)

        tree_scroll_y = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.city_tree.yview)
        tree_scroll_x = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=self.city_tree.xview)
        self.city_tree.configure(yscrollcommand=tree_scroll_y.set, xscrollcommand=tree_scroll_x.set)

        self.city_tree.grid(row=0, column=0, sticky="nsew")
        tree_scroll_y.grid(row=0, column=1, sticky="ns")
        tree_scroll_x.grid(row=1, column=0, sticky="ew")
        tree_frame.rowconfigure(0, weight=1)
        tree_frame.columnconfigure(0, weight=1)

        # 存储 item_id → adcode 映射
        self._city_item_map = {}

        # 点击树列切换复选框
        self.city_tree.bind("<ButtonRelease-1>", self._on_city_click)
        self.city_tree.bind("<space>", lambda e: self._city_toggle_selected())
        self.city_tree.bind("<Control-a>", lambda e: (self._city_select_all(), "break"))

        # 填充
        self._populate_city_tree()

    def _populate_city_tree(self, filter_text=""):
        self.city_tree.delete(*self.city_tree.get_children())
        self._city_item_map.clear()
        keyword = filter_text.strip().lower()
        self._insert_city_nodes(self.city_tree_data, "", keyword)
        # 默认展开省级
        if not keyword:
            for child in self.city_tree.get_children():
                self.city_tree.item(child, open=True)
        self._update_city_count_label()

    def _insert_city_nodes(self, nodes, parent, keyword):
        """递归插入城市树节点，keyword 为空则全部插入"""
        for code in sorted(nodes.keys()):
            node = nodes[code]
            name = node["name"]
            adcode = node["adcode"]
            children = node.get("children", {})

            match_self = (not keyword or
                          keyword in adcode.lower() or
                          keyword in name.lower())
            match_children = self._city_any_match(children, keyword) if keyword else True

            if not match_self and not match_children:
                continue

            mark = self._get_city_mark(adcode, children)
            display = f"{mark} {name}"
            item_id = self.city_tree.insert(parent, tk.END, text=display, values=(adcode,),
                                            open=(keyword != ""))
            self._city_item_map[item_id] = adcode

            if children:
                self._insert_city_nodes(children, item_id, keyword)
                # 更新父节点标记
                self._update_city_item_mark(item_id)

    def _city_any_match(self, nodes, keyword):
        if not keyword:
            return True
        for code, node in nodes.items():
            if (keyword in node["adcode"].lower() or
                keyword in node["name"].lower() or
                    self._city_any_match(node.get("children", {}), keyword)):
                return True
        return False

    def _get_city_mark(self, adcode, children):
        """计算节点复选框标记：☑ ☐ ◐"""
        if not children:
            return "☑" if adcode in self._selected_adcodes else "☐"

        all_child_adcodes = self._collect_descendant_adcodes(children)
        if not all_child_adcodes:
            return "☑" if adcode in self._selected_adcodes else "☐"

        selected_children = all_child_adcodes & self._selected_adcodes
        if len(selected_children) == len(all_child_adcodes):
            return "☑"
        elif len(selected_children) == 0:
            return "☐"
        else:
            return "◐"

    def _collect_descendant_adcodes(self, children):
        """收集某个子树中所有叶子节点的 adcode"""
        result = set()
        for code, node in children.items():
            sub = node.get("children", {})
            if sub:
                result |= self._collect_descendant_adcodes(sub)
            else:
                result.add(node["adcode"])
        # 同时采集中间节点的 adcode（中类/市级）
        for code, node in children.items():
            result.add(node["adcode"])
        return result

    def _update_city_item_mark(self, item_id):
        """根据 _selected_adcodes 更新单个节点的显示标记"""
        adcode = self._city_item_map.get(item_id)
        if not adcode:
            return

        # 找此节点的子节点
        children_map = self._find_children_map(adcode)
        mark = self._get_city_mark(adcode, children_map)
        current_text = self.city_tree.item(item_id, "text")
        # 替换第一个字符（标记）
        new_text = mark + current_text[1:]
        self.city_tree.item(item_id, text=new_text)

    def _find_children_map(self, adcode):
        """根据 adcode 在 city_tree_data 中查找对应的 children dict"""
        # 在树中搜索
        def search(nodes):
            for code, node in nodes.items():
                if node["adcode"] == adcode:
                    return node.get("children", {})
                result = search(node.get("children", {}))
                if result is not None:
                    return result
            return None

        result = search(self.city_tree_data)
        return result if result else {}

    def _city_toggle_node(self, item_id):
        """切换一个节点及其所有后代的选中状态"""
        adcode = self._city_item_map.get(item_id)
        if not adcode:
            return

        # 找到此节点下所有 adcode
        children_map = self._find_children_map(adcode)
        all_codes = {adcode} | self._collect_descendant_adcodes(children_map)

        # 判断新状态：如果当前全选则取消，否则全选
        currently_selected = all_codes & self._selected_adcodes
        if len(currently_selected) == len(all_codes):
            # 全部取消
            self._selected_adcodes -= all_codes
        else:
            # 全部勾选
            self._selected_adcodes |= all_codes

        # 更新此节点及所有子孙的显示
        self._refresh_city_node_and_ancestors(item_id)

    def _refresh_city_node_and_ancestors(self, item_id):
        """刷新节点及其所有祖先的显示标记"""
        # 递归刷新子树
        self._refresh_subtree(item_id)
        # 向祖先传播
        parent = self.city_tree.parent(item_id)
        while parent:
            self._update_city_item_mark(parent)
            parent = self.city_tree.parent(parent)

    def _refresh_subtree(self, item_id):
        """递归刷新子树所有节点的标记"""
        self._update_city_item_mark(item_id)
        for child in self.city_tree.get_children(item_id):
            self._refresh_subtree(child)

    def _on_city_click(self, event):
        """点击切换复选框"""
        region = self.city_tree.identify_region(event.x, event.y)
        if region not in ("tree", "cell"):
            return
        col = self.city_tree.identify_column(event.x)
        # 只响应树列 (#0) 的点击
        if col != "#0":
            return
        item_id = self.city_tree.identify_row(event.y)
        if not item_id:
            return
        self._city_toggle_node(item_id)
        self._update_city_count_label()

    def _city_toggle_selected(self):
        """空格键切换选中行"""
        for item_id in self.city_tree.selection():
            self._city_toggle_node(item_id)
        self._update_city_count_label()

    def _apply_city_filter(self):
        self._populate_city_tree(self.city_filter_var.get())

    def _city_select_all(self):
        self._selected_adcodes = set(self._all_adcodes)
        self._populate_city_tree(self.city_filter_var.get().strip())

    def _city_select_none(self):
        self._selected_adcodes.clear()
        self._populate_city_tree(self.city_filter_var.get().strip())

    def _city_invert(self):
        self._selected_adcodes = set(self._all_adcodes) - self._selected_adcodes
        self._populate_city_tree(self.city_filter_var.get().strip())

    def _city_expand_all(self):
        def expand(parent=""):
            for child in self.city_tree.get_children(parent):
                self.city_tree.item(child, open=True)
                expand(child)
        expand()

    def _city_collapse_all(self):
        def collapse(parent=""):
            for child in self.city_tree.get_children(parent):
                self.city_tree.item(child, open=False)
                collapse(child)
        collapse()

    def _update_city_count_label(self):
        self.city_count_label.config(
            text=f"已选: {len(self._selected_adcodes)} / {len(self._all_adcodes)}")

    # ---------- 行政边界标签页 ----------

    def _build_district_tab(self):
        tab = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(tab, text="行政边界")

        # 上半部分：配置 + 城市树
        upper = ttk.Frame(tab)
        upper.pack(fill=tk.BOTH, expand=True)

        # 左侧配置
        left = ttk.LabelFrame(upper, text="参数配置", padding="5")
        left.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 10))

        ttk.Label(left, text="子层级 (0-3):").pack(anchor=tk.W, pady=(0, 3))
        self.district_level = ttk.Spinbox(left, from_=0, to=3, width=5)
        self.district_level.set("1")
        self.district_level.pack(anchor=tk.W, pady=(0, 10))

        ttk.Label(left, text="0=仅本级  1=下延一级\n2=下延两级  3=下延三级",
                  foreground="gray", font=("", 8)).pack(anchor=tk.W)

        # 右侧城市选择
        right = ttk.LabelFrame(upper, text="选择行政区", padding="5")
        right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        top_bar = ttk.Frame(right)
        top_bar.pack(fill=tk.X, pady=(0, 3))

        self.district_filter_var = tk.StringVar()
        self.district_filter_var.trace_add("write", lambda *a: self._populate_district_tree())
        ttk.Entry(top_bar, textvariable=self.district_filter_var, width=18).pack(side=tk.LEFT)
        ttk.Label(top_bar, text=" 筛选", foreground="gray").pack(side=tk.LEFT)
        self.district_count_label = ttk.Label(top_bar, text="")
        self.district_count_label.pack(side=tk.RIGHT)

        tree_fr = ttk.Frame(right)
        tree_fr.pack(fill=tk.BOTH, expand=True)

        self.district_tree = ttk.Treeview(tree_fr, columns=("adcode",), show="tree headings",
                                          selectmode="extended")
        self.district_tree.heading("#0", text="行政区")
        self.district_tree.heading("adcode", text="代码")
        self.district_tree.column("#0", width=280)
        self.district_tree.column("adcode", width=90, anchor=tk.CENTER)

        ds = ttk.Scrollbar(tree_fr, orient=tk.VERTICAL, command=self.district_tree.yview)
        self.district_tree.configure(yscrollcommand=ds.set)
        self.district_tree.grid(row=0, column=0, sticky="nsew")
        ds.grid(row=0, column=1, sticky="ns")
        tree_fr.rowconfigure(0, weight=1)
        tree_fr.columnconfigure(0, weight=1)

        self.district_tree.bind("<ButtonRelease-1>", self._on_district_click)

        self._district_items = {}
        self._district_selected = set()
        self._populate_district_tree()

        # 按钮栏 — 放在树和日志之间，始终可见
        btn_bar = ttk.Frame(tab)
        btn_bar.pack(fill=tk.X, pady=(6, 2))

        ttk.Button(btn_bar, text="🔍 抓取边界", command=self._on_fetch_district).pack(
            side=tk.LEFT, padx=(0, 5))
        ttk.Button(btn_bar, text="📦 导出 SHP + CSV", command=self._on_export_district).pack(
            side=tk.LEFT, padx=(0, 15))
        ttk.Label(btn_bar, text="提示：勾选行政区 → 抓取边界 → 导出",
                  foreground="gray").pack(side=tk.LEFT)

        # 日志
        lower = ttk.Frame(tab)
        lower.pack(fill=tk.BOTH, expand=True)

        self.district_log = tk.Text(lower, wrap=tk.WORD, state=tk.DISABLED,
                                    font=("Consolas", 9), height=6,
                                    bg="#FFFFFF", fg="#1E1B4B")
        dlog_scroll = ttk.Scrollbar(lower, command=self.district_log.yview)
        self.district_log.configure(yscrollcommand=dlog_scroll.set)
        self.district_log.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        dlog_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        self._district_data = []  # 存放抓取结果

    def _populate_district_tree(self, filter_text=""):
        self.district_tree.delete(*self.district_tree.get_children())
        self._district_items.clear()
        keyword = filter_text.strip().lower()

        def insert_nodes(nodes, parent):
            for code in sorted(nodes.keys()):
                node = nodes[code]
                name = node["name"]
                adcode = node["adcode"]
                children = node.get("children", {})

                match = (not keyword or
                         keyword in adcode.lower() or
                         keyword in name.lower() or
                         self._dist_any_match(children, keyword))

                if not match:
                    continue

                mark = "☑" if adcode in self._district_selected else "☐"
                display = f"{mark} {name}"
                item = self.district_tree.insert(parent, tk.END, text=display,
                                                 values=(adcode,),
                                                 open=(not keyword))
                self._district_items[item] = adcode

                if children:
                    insert_nodes(children, item)

        insert_nodes(self.city_tree_data, "")
        self._update_district_count_label()

    def _dist_any_match(self, nodes, keyword):
        if not keyword:
            return True
        for code, node in nodes.items():
            if (keyword in node["adcode"].lower() or
                keyword in node["name"].lower() or
                    self._dist_any_match(node.get("children", {}), keyword)):
                return True
        return False

    def _on_district_click(self, event):
        region = self.district_tree.identify_region(event.x, event.y)
        if region not in ("tree", "cell"):
            return
        col = self.district_tree.identify_column(event.x)
        if col != "#0":
            return
        item = self.district_tree.identify_row(event.y)
        if not item:
            return
        adcode = self._district_items.get(item)
        if not adcode:
            return

        if adcode in self._district_selected:
            self._district_selected.discard(adcode)
        else:
            self._district_selected.add(adcode)

        current = self.district_tree.item(item, "text")
        mark = "☑" if adcode in self._district_selected else "☐"
        self.district_tree.item(item, text=mark + current[1:])
        self._update_district_count_label()

    def _update_district_count_label(self):
        self.district_count_label.config(
            text=f"已选: {len(self._district_selected)}")

    def _dlog(self, msg):
        ts = datetime.now().strftime("%H:%M:%S")
        self.district_log.configure(state=tk.NORMAL)
        self.district_log.insert(tk.END, f"[{ts}] {msg}\n")
        self.district_log.see(tk.END)
        self.district_log.configure(state=tk.DISABLED)

    def _parse_district_polyline(self, polyline_str):
        """解析高德行政边界 polyline 字符串 → Polygon/MultiPolygon"""
        if not polyline_str:
            return None
        polygons = []
        for part in polyline_str.split("|"):
            coords = []
            for pair in part.split(";"):
                if "," in pair:
                    lng, lat = pair.split(",")
                    try:
                        coords.append((float(lng), float(lat)))
                    except ValueError:
                        pass
            if len(coords) >= 3:
                poly = Polygon(coords)
                if poly.is_valid and not poly.is_empty:
                    polygons.append(poly)
        if not polygons:
            return None
        return MultiPolygon(polygons) if len(polygons) > 1 else polygons[0]

    def _on_fetch_district(self):
        api_key = self.entry_api_key.get().strip()
        if not api_key:
            messagebox.showwarning("提示", "请先在「API 配置」页填写高德 API Key")
            return
        if not self._district_selected:
            messagebox.showwarning("提示", "请勾选至少一个行政区")
            return

        level = int(self.district_level.get())
        self._district_data = []
        self.district_log.configure(state=tk.NORMAL)
        self.district_log.delete("1.0", tk.END)
        self.district_log.configure(state=tk.DISABLED)

        total = len(self._district_selected)

        # 构建 adcode→name 映射
        city_map = {c["adcode"]: c["city"] for c in self.cities}

        for i, adcode in enumerate(self._district_selected):
            name = city_map.get(adcode, adcode)
            self._dlog(f"[{i+1}/{total}] 获取: {name} ({adcode})")

            url = (
                "https://restapi.amap.com/v3/config/district"
                f"?key={api_key}"
                f"&keywords={adcode}"
                f"&subdistrict={level}"
                f"&extensions=all"
            )
            try:
                resp = requests.get(url, timeout=30)
                data = json.loads(resp.content)
            except Exception as e:
                self._dlog(f"  请求失败: {e}")
                continue

            if data.get("status") != "1":
                self._dlog(f"  API 错误: {data.get('info', '未知')}")
                continue

            districts = data.get("districts", [])
            if not districts:
                self._dlog("  无数据")
                continue

            self._collect_districts(districts, level, name, adcode)

            time.sleep(0.3)

        self._dlog(f"完成！共获取 {len(self._district_data)} 个行政区划")
        messagebox.showinfo("完成", f"共获取 {len(self._district_data)} 个行政区划\n点击「导出 SHP + CSV」保存")

    def _collect_districts(self, districts, max_level, parent_name, parent_code, current_level=0):
        """递归收集 district API 返回的行政区数据"""
        for d in districts:
            d_name = d.get("name", "")
            d_adcode = d.get("adcode", "")
            d_level = d.get("level", "")
            polyline = d.get("polyline", "")

            geom = self._parse_district_polyline(polyline)

            self._district_data.append({
                "名称": d_name,
                "编码": d_adcode,
                "层级": d_level,
                "上级名称": parent_name,
                "上级编码": parent_code,
                "geometry": geom,
            })

            if geom:
                self._dlog(f"  {d_name} ({d_adcode}) [{d_level}] — 有边界")
            else:
                self._dlog(f"  {d_name} ({d_adcode}) [{d_level}] — 无边界")

            # 递归子级
            sub = d.get("districts", [])
            if sub and current_level < max_level:
                self._collect_districts(sub, max_level, d_name, d_adcode, current_level + 1)

    def _on_export_district(self):
        if not self._district_data:
            messagebox.showwarning("提示", "没有数据，请先「抓取边界」")
            return

        geo_rows = [r for r in self._district_data if r.get("geometry") is not None]
        if not geo_rows:
            messagebox.showwarning("提示", "没有包含边界的行政区数据")
            return

        default_name = f"district_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        shp_path = filedialog.asksaveasfilename(
            title="导出行政边界 Shapefile",
            defaultextension=".shp",
            filetypes=[("Shapefile", "*.shp"), ("All files", "*.*")],
            initialfile=f"{default_name}.shp",
        )
        if not shp_path:
            return

        try:
            w = shp.Writer(shp_path, shapeType=shp.POLYGON, encoding="utf-8")
            w.field("名称", "C", "100")
            w.field("编码", "C", "20")
            w.field("层级", "C", "20")

            csv_rows = []
            for r in geo_rows:
                geom = r.get("geometry")
                if geom is None:
                    continue
                # 处理 Polygon / MultiPolygon → pyshp 格式
                if geom.geom_type == "Polygon":
                    exterior = [(c[0], c[1]) for c in geom.exterior.coords]
                    w.poly([exterior])
                elif geom.geom_type == "MultiPolygon":
                    parts = []
                    for poly in geom.geoms:
                        parts.append([(c[0], c[1]) for c in poly.exterior.coords])
                    if parts:
                        w.poly(parts)
                else:
                    continue
                w.record(r.get("名称", ""), r.get("编码", ""), r.get("层级", ""))
                csv_rows.append({"名称": r.get("名称", ""), "编码": r.get("编码", ""),
                                 "层级": r.get("层级", ""), "上级名称": r.get("上级名称", ""),
                                 "上级编码": r.get("上级编码", "")})
            w.close()

            # PRJ
            prj_path = shp_path.replace(".shp", ".prj")
            with open(prj_path, "w", encoding="utf-8") as f:
                f.write('GEOGCS["WGS 84",DATUM["WGS_1984",SPHEROID["WGS 84",6378137,298.257223563]],PRIMEM["Greenwich",0],UNIT["degree",0.0174532925199433]]')

            # CSV
            csv_path = shp_path.replace(".shp", ".csv")
            import csv
            if csv_rows:
                with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
                    writer = csv.DictWriter(f, fieldnames=list(csv_rows[0].keys()))
                    writer.writeheader()
                    writer.writerows(csv_rows)

            self._dlog(f"SHP: {shp_path} ({len(csv_rows)} 条)")
            self._dlog(f"CSV: {csv_path}")
            messagebox.showinfo("导出完成",
                                f"{len(csv_rows)} 个行政区划\n\n"
                                f"SHP: {shp_path}\nCSV: {csv_path}")

        except Exception as e:
            import traceback
            self._dlog(f"导出失败: {e}\n{traceback.format_exc()}")
            messagebox.showerror("导出失败", str(e))

    # ---------- 操作按钮 ----------

    def _get_config_from_ui(self):
        return {
            "api_key": self.entry_api_key.get().strip(),
            "poi_types": self.entry_poi_types.get().strip(),
            "page_size": int(self.spin_page_size.get()),
            "db_host": self.entry_db_host.get().strip(),
            "db_port": int(self.entry_db_port.get()),
            "db_user": self.entry_db_user.get().strip(),
            "db_password": self.entry_db_password.get(),
            "db_name": self.entry_db_name.get().strip(),
            "table_name": self.entry_table_name.get().strip(),
            "selected_cities": list(self._selected_adcodes),
        }

    def _load_config_to_ui(self):
        c = self.config_mgr.config
        self.entry_api_key.insert(0, c.get("api_key", ""))
        self.entry_poi_types.insert(0, c.get("poi_types", ""))
        self.spin_page_size.set(str(c.get("page_size", 20)))
        self.entry_db_host.insert(0, c.get("db_host", "localhost"))
        self.entry_db_port.insert(0, str(c.get("db_port", 5432)))
        self.entry_db_user.insert(0, c.get("db_user", "postgres"))
        self.entry_db_password.insert(0, c.get("db_password", ""))
        self.entry_db_name.insert(0, c.get("db_name", "pgadmind"))
        self.entry_table_name.insert(0, c.get("table_name", "gaode_poi"))

    def _on_save_config(self):
        self.config_mgr.config = self._get_config_from_ui()
        self.config_mgr.save()
        self._log("配置已保存到 config.json")

    def _on_browse_poi(self):
        if not self.poi_tree:
            messagebox.showinfo("提示", "未找到 amap_poicode.xlsx\n请手动输入 POI 类别编码（如 061205）")
            return
        dlg = PoiTypeDialog(self, self.poi_tree, self.entry_poi_types.get().strip())
        if dlg.result:
            self.entry_poi_types.delete(0, tk.END)
            self.entry_poi_types.insert(0, dlg.result)

    def _on_test_db(self):
        cfg = self._get_config_from_ui()
        try:
            conn = psycopg2.connect(
                host=cfg["db_host"], port=cfg["db_port"],
                user=cfg["db_user"], password=cfg["db_password"],
                dbname=cfg["db_name"], client_encoding="UTF8",
            )
            conn.close()
            messagebox.showinfo("成功", "数据库连接成功！")
            self._log("数据库连接测试: 成功")
        except Exception as e:
            messagebox.showerror("失败", f"连接失败:\n{e}")
            self._log(f"数据库连接测试: 失败 - {e}")

    def _on_create_table(self):
        cfg = self._get_config_from_ui()
        try:
            conn = psycopg2.connect(
                host=cfg["db_host"], port=cfg["db_port"],
                user=cfg["db_user"], password=cfg["db_password"],
                dbname=cfg["db_name"], client_encoding="UTF8",
            )
            sql = f'CREATE TABLE IF NOT EXISTS "{cfg["table_name"]}" ({TABLE_COLUMNS_DEF})'
            with conn.cursor() as cur:
                cur.execute(sql)
            conn.commit()
            conn.close()
            messagebox.showinfo("成功", f"表 \"{cfg['table_name']}\" 已就绪")
            self._log(f"数据表 \"{cfg['table_name']}\" 创建/确认成功")
        except Exception as e:
            messagebox.showerror("失败", f"创建表失败:\n{e}")
            self._log(f"创建表失败: {e}")

    def _on_export_shp(self):
        """导出数据库 POI 表为点 Shapefile — 可选择表"""
        cfg = self._get_config_from_ui()

        # 先获取数据库中所有 POI 表列表
        try:
            conn = psycopg2.connect(
                host=cfg["db_host"], port=cfg["db_port"],
                user=cfg["db_user"], password=cfg["db_password"],
                dbname=cfg["db_name"], client_encoding="UTF8",
            )
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT DISTINCT t.table_name
                    FROM information_schema.tables t
                    JOIN information_schema.columns c
                      ON t.table_name = c.table_name AND t.table_schema = c.table_schema
                    WHERE t.table_schema='public' AND t.table_type='BASE TABLE'
                      AND c.column_name = '编号'
                    ORDER BY t.table_name DESC
                """)
                tables = [r[0] for r in cur.fetchall()]
            conn.close()
        except Exception as e:
            messagebox.showerror("错误", f"获取表列表失败:\n{e}")
            return

        if not tables:
            messagebox.showinfo("提示", "数据库中还没有 POI 数据表，请先爬取数据。")
            return

        # 弹出表选择对话框
        dlg = tk.Toplevel(self)
        dlg.title("选择要导出的数据表")
        dlg.geometry("400x350")
        dlg.transient(self)
        dlg.grab_set()
        dlg.resizable(False, False)

        ttk.Label(dlg, text="请选择要导出的 POI 数据表：",
                  font=("Microsoft YaHei UI", 10)).pack(pady=(15, 10))

        list_frame = ttk.Frame(dlg)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=15)

        lb = tk.Listbox(list_frame, font=("Consolas", 10), selectmode=tk.SINGLE)
        lb_scroll = ttk.Scrollbar(list_frame, command=lb.yview)
        lb.configure(yscrollcommand=lb_scroll.set)
        lb.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        lb_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        for t in tables:
            lb.insert(tk.END, t)

        # 默认选中第一项
        if tables:
            lb.selection_set(0)
            lb.activate(0)

        chosen_table = [None]

        def on_ok():
            sel = lb.curselection()
            if sel:
                chosen_table[0] = lb.get(sel[0])
            dlg.destroy()

        btn_frame = ttk.Frame(dlg)
        btn_frame.pack(fill=tk.X, padx=15, pady=(10, 15))
        ttk.Button(btn_frame, text="确定", command=on_ok).pack(side=tk.RIGHT, padx=(5, 0))
        ttk.Button(btn_frame, text="取消", command=dlg.destroy).pack(side=tk.RIGHT)

        # 双击选择
        lb.bind("<Double-1>", lambda e: on_ok())
        # Enter 键
        dlg.bind("<Return>", lambda e: on_ok())

        dlg.wait_window()

        table = chosen_table[0]
        if not table:
            return

        default_name = f"{table}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        shp_path = filedialog.asksaveasfilename(
            title=f"导出 {table} → Shapefile",
            defaultextension=".shp",
            filetypes=[("Shapefile", "*.shp"), ("All files", "*.*")],
            initialfile=f"{default_name}.shp",
        )
        if not shp_path:
            return

        try:
            conn = psycopg2.connect(
                host=cfg["db_host"], port=cfg["db_port"],
                user=cfg["db_user"], password=cfg["db_password"],
                dbname=cfg["db_name"], client_encoding="UTF8",
            )

            with conn.cursor() as cur:
                cur.execute(f'SELECT COUNT(*) FROM "{table}"')
                total = cur.fetchone()[0]

            if total == 0:
                messagebox.showinfo("提示", f"表 \"{table}\" 中无数据")
                conn.close()
                return

            self._log(f"正在导出表 [{table}] 共 {total} 条记录...")
            self.status_label.config(text=f"导出 {table}: {total} 条...")

            with conn.cursor() as cur:
                cur.execute(f'SELECT * FROM "{table}"')
                columns = [desc[0] for desc in cur.description]
                rows = []
                for row in cur:
                    rows.append(dict(zip(columns, row)))
            conn.close()

            # 收集有坐标的行
            valid = []
            for r in rows:
                loc = r.get("经纬度", "")
                if loc and "," in loc:
                    try:
                        lng_str, lat_str = loc.split(",")
                        lng, lat = float(lng_str), float(lat_str)
                        valid.append((lng, lat, r))
                    except (ValueError, TypeError):
                        pass

            if not valid:
                messagebox.showinfo("提示", "没有包含有效坐标的记录")
                return

            # 用 pyshp 写 SHP
            w = shp.Writer(shp_path, shapeType=shp.POINT, encoding="utf-8")
            # 字段：用前 10 个字段名，限制长度
            field_names = [c for c in columns if c != "经纬度"][:30]
            for name in field_names:
                w.field(name[:10], "C", "254")

            csv_rows = []
            for lng, lat, r in valid:
                w.point(lng, lat)
                rec = [str(r.get(n, "") or "")[:254] for n in field_names]
                w.record(*rec)
                # CSV 行
                csv_row = {n: r.get(n, "") for n in columns}
                csv_row["_lng"] = lng
                csv_row["_lat"] = lat
                csv_rows.append(csv_row)
            w.close()

            # 写 PRJ 文件
            prj_path = shp_path.replace(".shp", ".prj")
            with open(prj_path, "w", encoding="utf-8") as f:
                f.write('GEOGCS["WGS 84",DATUM["WGS_1984",SPHEROID["WGS 84",6378137,298.257223563]],PRIMEM["Greenwich",0],UNIT["degree",0.0174532925199433]]')

            # CSV
            csv_path = shp_path.replace(".shp", ".csv")
            import csv
            if csv_rows:
                with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
                    writer = csv.DictWriter(f, fieldnames=list(csv_rows[0].keys()))
                    writer.writeheader()
                    writer.writerows(csv_rows)

            self._log(f"导出完成: {len(valid)} 条 → {shp_path}")
            self._log(f"CSV: {csv_path}")
            self.status_label.config(text=f"导出完成: {len(valid)} 条")
            messagebox.showinfo("导出完成",
                                f"{len(valid)} 条记录\n\n"
                                f"SHP: {shp_path}\nCSV: {csv_path}\n\n"
                                "坐标系: WGS84 (EPSG:4326)")

        except Exception as e:
            import traceback
            self._log(f"导出失败: {e}\n{traceback.format_exc()}")
            messagebox.showerror("导出失败", str(e))

    def _on_start(self):
        self.config_mgr.config = self._get_config_from_ui()
        self.config_mgr.save()

        if not self.config_mgr.get("api_key"):
            messagebox.showwarning("提示", "请填写高德 API Key")
            return

        if not self._selected_adcodes:
            messagebox.showwarning("提示", "请至少选择一个城市")
            return

        # 智能生成表名：地域 + POI类型
        current_table = self.entry_table_name.get().strip()
        default_table = "gaode_poi"
        if current_table == default_table or not current_table:
            # 查找 POI 类型名称
            type_name = ""
            type_code = self.entry_poi_types.get().strip()
            if type_code and self.poi_tree:
                def find_name(tree, target):
                    for code, node in tree.items():
                        if code == target:
                            return node["name"]
                        result = find_name(node.get("children", {}), target)
                        if result:
                            return result
                    return None
                type_name = find_name(self.poi_tree, type_code) or ""
            # 简化POI类别名（去掉过长的层级描述）
            if " > " in type_name:
                type_name = type_name.split(" > ")[-1]
            if not type_name:
                type_name = type_code

            # 确定地域范围
            selected_cities = [c for c in self.cities if c["adcode"] in self._selected_adcodes]
            prov_codes = set(c["adcode"][:2] for c in selected_cities)

            if len(prov_codes) == 1:
                # 同一省份，找省份名
                prov_code = list(prov_codes)[0]
                prov_name = ""
                for code, node in self.city_tree_data.items():
                    if code == prov_code:
                        prov_name = node["name"]
                        break
                # 去掉"省"/"市"后缀避免冗余
                for suffix in ["省", "市", "自治区", "维吾尔自治区", "回族自治区", "壮族自治区"]:
                    if prov_name.endswith(suffix) and len(prov_name) > len(suffix) + 1:
                        prov_name = prov_name[:-len(suffix)]
                        break
                # 判断是全省还是部分城市
                prov_cities = set(c["adcode"] for c in selected_cities)
                prov_all = set(c["adcode"] for c in self.cities if c["adcode"].startswith(prov_code))
                if prov_cities == prov_all or len(selected_cities) >= len(prov_all) * 0.8:
                    region = prov_name
                elif len(selected_cities) == 1:
                    region = selected_cities[0]["city"]
                else:
                    region = f"{prov_name}{len(selected_cities)}城"
            else:
                region = f"多省{len(prov_codes)}"

            new_table = f"poi_{region}_{type_name}"
            # 清理表名中的特殊字符
            new_table = new_table.replace(" ", "").replace("/", "").replace("\\", "")
            if len(new_table) > 50:
                new_table = new_table[:50]
            self.entry_table_name.delete(0, tk.END)
            self.entry_table_name.insert(0, new_table)
            self.config_mgr.config["table_name"] = new_table
            self.config_mgr.save()

        selected_list = [c for c in self.cities if c["adcode"] in self._selected_adcodes]
        tbl = self.entry_table_name.get().strip()

        ok = messagebox.askokcancel(
            "开始爬取",
            f"即将爬取 {len(selected_list)} 个城市\n"
            f"POI 类别: {self.config_mgr.get('poi_types')}\n"
            f"数据表: {tbl}\n"
            f"数据库: {self.config_mgr.get('db_name')}\n\n"
            f"确认开始？"
        )
        if not ok:
            return

        self.log_text.configure(state=tk.NORMAL)
        self.log_text.delete("1.0", tk.END)
        self.log_text.configure(state=tk.DISABLED)

        self.progress_var.set(0)
        self.sub_progress_var.set(0)
        self.status_label.config(text="正在启动...")

        self.stop_event.clear()
        self.current_batch_id = datetime.now().strftime("%Y%m%d_%H%M%S")
        self._log(f"抓取批次: {self.current_batch_id}")
        self.crawler = CrawlerEngine(
            self.config_mgr.config,
            selected_list,
            self.log_queue,
            self.stop_event,
            self.current_batch_id,
        )
        self.crawler.start()

        self.btn_start.config(state=tk.DISABLED)
        self.btn_stop.config(state=tk.NORMAL)
        self._log("开始爬取...")

    def _on_stop(self):
        if messagebox.askyesno("确认", "确定要停止爬取吗？\n当前正在进行的请求会完成后再停止。"):
            self._log("正在停止（等待当前请求完成）...")
            self.stop_event.set()
            self.btn_stop.config(state=tk.DISABLED)

    def _on_close(self):
        if self.crawler and self.crawler.is_alive():
            if messagebox.askyesno("退出", "爬虫正在运行，确定要退出吗？"):
                self.stop_event.set()
                self.destroy()
        else:
            self.destroy()

    # ---------- 日志与进度轮询 ----------

    def _log(self, msg):
        ts = datetime.now().strftime("%H:%M:%S")
        self.log_text.configure(state=tk.NORMAL)
        self.log_text.insert(tk.END, f"[{ts}] {msg}\n")
        self.log_text.see(tk.END)
        self.log_text.configure(state=tk.DISABLED)

    def _poll_log(self):
        try:
            while True:
                msg = self.log_queue.get_nowait()
                if isinstance(msg, tuple):
                    mtype = msg[0]
                    if mtype == "log":
                        self._log(msg[1])
                    elif mtype == "error":
                        self._log(f"错误: {msg[1]}")
                    elif mtype == "progress_city":
                        _, idx, total, name = msg
                        pct = (idx / total) * 100 if total > 0 else 0
                        self.progress_var.set(pct)
                        self.status_label.config(text=f"正在爬取: {name} ({idx + 1}/{total})")
                        self.sub_progress_var.set(0)
                    elif mtype == "progress_page":
                        _, page, total = msg
                        pct = (page / total) * 100 if total > 0 else 0
                        self.sub_progress_var.set(pct)
                    elif mtype == "done":
                        self.progress_var.set(100)
                        self.sub_progress_var.set(100)
                        self.status_label.config(text=f"完成！共写入 {msg[1]} 条")
                        self.btn_start.config(state=tk.NORMAL)
                        self.btn_stop.config(state=tk.DISABLED)
                else:
                    self._log(str(msg))
        except queue.Empty:
            pass

        if self.crawler and not self.crawler.is_alive():
            if str(self.btn_start["state"]) == tk.DISABLED:
                status = self.status_label.cget("text")
                if status not in ("就绪",) and not status.startswith("完成"):
                    self.btn_start.config(state=tk.NORMAL)
                    self.btn_stop.config(state=tk.DISABLED)
                    if self.status_label.cget("text") == "正在启动...":
                        self.status_label.config(text="已停止")

        self.after(100, self._poll_log)


def show_splash():
    """绚丽开场动画"""
    splash = tk.Tk()
    splash.overrideredirect(True)
    splash.attributes("-topmost", True)
    sw, sh = 560, 380
    sx = (splash.winfo_screenwidth() - sw) // 2
    sy = (splash.winfo_screenheight() - sh) // 2
    splash.geometry(f"{sw}x{sh}+{sx}+{sy}")
    splash.configure(bg="#0A0A1A")
    splash.attributes("-alpha", 0.0)

    # 画布
    canvas = tk.Canvas(splash, width=sw, height=sh, bg="#0A0A1A",
                       highlightthickness=0)
    canvas.pack(fill=tk.BOTH, expand=True)

    # 背景圆形光晕
    colors = ["#1A1040", "#1E1145", "#221250", "#1A1040", "#0E0A28"]
    circles = []
    for i, (cx, cy, r, color) in enumerate([
        (280, 140, 180, "#1A1040"), (180, 240, 120, "#221250"),
        (400, 100, 90, "#1E1145"), (100, 80, 60, "#2D1A60"),
        (460, 260, 70, "#251850"),
    ]):
        c = canvas.create_oval(cx-r, cy-r, cx+r, cy+r, fill=color,
                                outline="", tags=f"bg{i}")
        circles.append(c)

    # 脉冲环
    ring1 = canvas.create_oval(100, 70, 460, 310, outline="#6C3CE0",
                                width=2, dash=(4, 8), tags="ring")
    ring2 = canvas.create_oval(120, 90, 440, 290, outline="#8B5CF6",
                                width=1, dash=(2, 16), tags="ring")

    # 主标题
    title = canvas.create_text(280, 130, text="高德地图 POI 爬虫工具",
                                fill="#8B5CF6", font=("Microsoft YaHei UI", 26, "bold"),
                                tags="title")
    subtitle = canvas.create_text(280, 170, text="Gaode POI Spider Pro",
                                   fill="#A78BFA", font=("Consolas", 14), tags="sub")

    # 装饰线
    canvas.create_line(160, 195, 400, 195, fill="#6C3CE0", width=1, tags="line")
    canvas.create_line(190, 200, 370, 200, fill="#8B5CF6", width=2, tags="line")

    # 开发者信息
    dev = canvas.create_text(280, 240, text="开发者：白栋博",
                              fill="#C4B5FD", font=("Microsoft YaHei UI", 13, "bold"),
                              tags="dev")
    wx = canvas.create_text(280, 270, text="微信：bdb15896810691",
                             fill="#A78BFA", font=("Microsoft YaHei UI", 12), tags="wx")

    # 底部提示
    loading = canvas.create_text(280, 330, text="● 正在启动...",
                                  fill="#6C3CE0", font=("Microsoft YaHei UI", 9), tags="load")

    # 粒子
    particles = []
    import random, math
    for _ in range(30):
        x = random.randint(20, sw-20)
        y = random.randint(20, sh-20)
        r = random.randint(1, 3)
        p = canvas.create_oval(x-r, y-r, x+r, y+r,
                                fill=random.choice(["#8B5CF6","#A78BFA","#6C3CE0","#C4B5FD"]),
                                outline="", tags="particle")
        particles.append((p, x, y, random.uniform(0, 2*math.pi), random.uniform(0.3, 1.2)))

    # 淡入
    for a in [0.1, 0.2, 0.35, 0.5, 0.7, 0.85, 0.95, 1.0]:
        splash.attributes("-alpha", a)
        splash.update()
        time.sleep(0.03)

    # 动画循环
    start_t = time.time()
    angle = 0
    while time.time() - start_t < 10:
        angle += 0.05
        # 脉冲环旋转效果
        canvas.itemconfig(ring1, dash=(4+int(math.sin(angle)*2), 8))
        canvas.itemconfig(ring2, dash=(2, 16+int(math.sin(angle*1.5)*4)))
        # 粒子移动
        for i, (p, px, py, pa, ps) in enumerate(particles):
            nx = px + math.cos(pa + angle*ps) * 15 * math.sin(angle*0.7)
            ny = py + math.sin(pa + angle*ps) * 15 * math.cos(angle*0.7)
            canvas.coords(p, nx-1, ny-1, nx+1, ny+1)
        # 标题呼吸
        scale = 1 + math.sin(angle*1.3) * 0.02
        canvas.itemconfig(title, font=("Microsoft YaHei UI", int(26*scale), "bold"))
        # loading 闪烁
        dots = "●" * (1 + int((time.time()*3) % 4))
        canvas.itemconfig(loading, text=f"{dots} 正在启动...")

        splash.update()
        time.sleep(0.02)

    # 淡出
    for a in [0.9, 0.7, 0.5, 0.3, 0.15, 0.05, 0.0]:
        splash.attributes("-alpha", a)
        splash.update()
        time.sleep(0.02)

    splash.destroy()


def main():
    import traceback
    def _show_error(exc_type, exc_val, exc_tb):
        msg = "".join(traceback.format_exception(exc_type, exc_val, exc_tb))
        try:
            tk.messagebox.showerror("程序异常", f"请将此信息反馈给开发者：\n\n{msg[-2000:]}")
        except Exception:
            pass
    tk.Tk.report_callback_exception = _show_error

    show_splash()

    try:
        app = App()
        app.mainloop()
    except Exception as e:
        traceback.print_exc()
        try:
            tk.messagebox.showerror("启动失败", str(e))
        except Exception:
            pass


if __name__ == "__main__":
    main()
